from pythoncom import com_error
import robot.libraries.Screenshot as screenshot
import os
from robot.api import logger
import  json
from PIL import Image
import pandas as pd
import json
from fpdf import FPDF
from PIL import Image
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import shutil
import pandas as pd
import datetime
from collections import defaultdict
from datetime import datetime
import calendar
import re


class Replay_Library:    
    def __init__(self):
        self.name = "DefaultReport"
    def copy_images(self, source_dir, target_dir):
        # Ensure the target directory exists
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)                       
        image_formats = ('.jpeg', '.jpg', '.png', '.gif', '.bmp', '.tiff', '.webp')
        for file_name in os.listdir(source_dir):
            file_path = os.path.join(source_dir, file_name)
            if os.path.isfile(file_path):
                try:
                    with Image.open(file_path) as img:
                        if file_name.lower().endswith(image_formats):
                            target_path = os.path.join(target_dir, file_name)
                            shutil.copy(file_path, target_path)
                            print(f"Copied: {file_name}")
                except Exception as e:
                    print(f"Skipped: {file_name} - Not a valid image file or format not supported. Error: {e}")
                    
    def remove_rows_before_start_row(self, file_path, sheet_name, start_row):
        start_row = int(start_row)
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        for _ in range(start_row - 1):  # Start row number is 1-based
            sheet.delete_rows(1)  # Always delete the first row
        workbook.save(file_path)
        workbook.close()
        return f"All rows before row {start_row} have been removed."

    def clean_excel(self, file_path, sheet_name):
        try:
            workbook = load_workbook(file_path)
            if sheet_name not in workbook.sheetnames:
                raise Exception(f"Sheet '{sheet_name}' does not exist in the file.")
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = cell.value.strip()  # Remove leading and trailing whitespace
            for row_idx in range(sheet.max_row, 0, -1):
                if all(sheet.cell(row=row_idx, column=col_idx).value in [None, ""] for col_idx in range(1, sheet.max_column + 1)):
                    sheet.delete_rows(row_idx)
            for col_idx in range(sheet.max_column, 0, -1):
                if all(sheet.cell(row=row_idx, column=col_idx).value in [None, ""] for row_idx in range(1, sheet.max_row + 1)):
                    sheet.delete_cols(col_idx)
            workbook.save(file_path)
            print("\033[92m❗ Excel sheet cleaned successfully and saved at:", file_path)
        except Exception as e:
            print("\033[92m❗ Failed to clean Excel sheet:", str(e))

    def fiori_html_report(self, excel_file, html_file, highlight_text):
        df = pd.read_excel(excel_file)
        # Initialize counters
        pass_count = 0
        warn_count = 0
        for index, row in df.iterrows():
            row_text = " ".join(row.astype(str).tolist())
            if 'Authorization check successful' in row_text:
                pass_count += 1
            if 'No authorization in user master record' in row_text:
                warn_count += 1
        current_time = datetime.datetime.utcnow().strftime('%d-%m-%Y Time: %H:%M:%S UTC')
        #HTML
        html_content = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Fiori LaunchPad Report</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 0;
                    background-color: #f4f4f4;
                }}
                .container {{
                    width: 90%;
                    margin: 30px auto;
                    padding: 20px;
                    background-color: #ffffff;
                    border-radius: 8px;
                    box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
                }}
                h1 {{
                    color: #1f3a64;
                    text-align: center;
                }}
                .count-box {{
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    gap: 20px;
                    margin: 20px 0;
                }}
                .count-item {{
                    display: flex;
                    align-items: center;
                    gap: 10px;
                    font-size: 18px;
                }}
                .pass-circle {{
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    background-color: green;
                    border-radius: 50%;
                    margin-right: 8px; /* Adds spacing between the circle and text */
                }}
                .warn-circle {{
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    background-color: red;
                    border-radius: 50%;
                    margin-right: 8px; /* Adds spacing between the circle and text */
                }}

                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th, td {{
                    padding: 12px;
                    text-align: left;
                    border: 1px solid #ddd;
                }}
                th {{
                    background-color: #1f3a64;
                    color: white;
                }}
                td {{
                    background-color: #fafafa;
                }}
                table tr:nth-child(even) td {{
                    background-color: #f2f2f2;
                }}
                table tr:hover td {{
                    background-color: #e9e9e9;
                }}
                .row-green {{
                    background-color: #d4edda;
                    color: green;
                }}
                .row-red {{
                    background-color: #ff0000 ;
                    color:red ;
                }}
            </style>
        </head>
        <body>
        <div class="container">
            <h1>Fiori LaunchPad Report</h1>
                <div class="count-box">
                    <div class="count-item">
                        <span class="pass-circle"></span> Authorization Count: <span>{pass_count}</span>
                    </div>
                    <div class="count-item">
                        <span class="warn-circle"></span> No Authorization Count: <span>{warn_count}</span>
                    </div>
                    <div class="count-item">
                        Generated At: <span>{current_time}</span>
                    </div>
                </div>
            <table>
                <thead>
                    <tr>
        """
        # Add table headers (columns from DataFrame)
        headers = df.columns.tolist()
        for header in headers:
            html_content += f"<th>{header}</th>"
        html_content += """
                    </tr>
                </thead>
                <tbody>
        """
        # Add table rows (data from DataFrame)
        for index, row in df.iterrows():
            # Check if the target text is in any cell of the row
            if row.astype(str).str.contains(highlight_text, case=False).any():
                html_content += '<tr class="row-red">'  
            else:
                html_content += '<tr class="row-green">'  
            for value in row:
                html_content += f"<td>{value}</td>"
            html_content += "</tr>"
        html_content += """
                </tbody>
            </table>
        </div>
        </body>
        </html>
        """
        with open(html_file, "w") as file:
            file.write(html_content)
        print(f"HTML report created successfully: {html_file}")

    def sarole_html_report(self, excel_file, html_file, highlight_text):
        df = pd.read_excel(excel_file)
        # Initialize counters
        pass_count = 0
        warn_count = 0
        for index, row in df.iterrows():
            row_text = " ".join(row.astype(str).tolist())
            if 'Authorization check successful' in row_text:
                pass_count += 1
            if 'No authorization in user master record' in row_text:
                warn_count += 1
        current_time = datetime.datetime.utcnow().strftime('%d-%m-%Y Time: %H:%M:%S UTC')
        #HTML
        html_content = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Security Authorization Report</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 0;
                    background-color: #f4f4f4;
                }}
                .container {{
                    width: 90%;
                    margin: 30px auto;
                    padding: 20px;
                    background-color: #ffffff;
                    border-radius: 8px;
                    box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
                }}
                h1 {{
                    color: #1f3a64;
                    text-align: center;
                }}
                .count-box {{
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    gap: 20px;
                    margin: 20px 0;
                }}
                .count-item {{
                    display: flex;
                    align-items: center;
                    gap: 10px;
                    font-size: 18px;
                }}
                .pass-circle {{
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    background-color: green;
                    border-radius: 50%;
                    margin-right: 8px; /* Adds spacing between the circle and text */
                }}
                .warn-circle {{
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    background-color: red;
                    border-radius: 50%;
                    margin-right: 8px; /* Adds spacing between the circle and text */
                }}

                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th, td {{
                    padding: 12px;
                    text-align: left;
                    border: 1px solid #ddd;
                }}
                th {{
                    background-color: #1f3a64;
                    color: white;
                }}
                td {{
                    background-color: #fafafa;
                }}
                table tr:nth-child(even) td {{
                    background-color: #f2f2f2;
                }}
                table tr:hover td {{
                    background-color: #e9e9e9;
                }}
                .row-green {{
                    background-color: #d4edda;
                    color: green;
                }}
                .row-red {{
                    background-color: #ff0000 ;
                    color:red ;
                }}
            </style>
        </head>
        <body>

        <div class="container">
            <h1>Security Authorization Report</h1>
                <div class="count-box">
                    <div class="count-item">
                        <span class="pass-circle"></span> Authorization Count: <span>{pass_count}</span>
                    </div>
                    <div class="count-item">
                        <span class="warn-circle"></span> No Authorization Count: <span>{warn_count}</span>
                    </div>
                    <div class="count-item">
                        Generated At: <span>{current_time}</span>
                    </div>
                </div>
            <table>
                <thead>
                    <tr>
        """
        # Add table headers (columns from DataFrame)
        headers = df.columns.tolist()
        for header in headers:
            html_content += f"<th>{header}</th>"
        html_content += """
                    </tr>
                </thead>
                <tbody>
        """
        # Add table rows (data from DataFrame)
        for index, row in df.iterrows():
            # Check if the target text is in any cell of the row
            if row.astype(str).str.contains(highlight_text, case=False).any():
                html_content += '<tr class="row-red">'  
            else:
                html_content += '<tr class="row-green">'  
            for value in row:
                html_content += f"<td>{value}</td>"
            html_content += "</tr>"
        html_content += """
                </tbody>
            </table>
        </div>

        </body>
        </html>
        """
        with open(html_file, "w") as file:
            file.write(html_content)
        print(f"HTML report created successfully: {html_file}")

    def excel_remove_multiple_columns(self, file_path, sheet_name, col_indices):
        col_indices = sorted([int(col) for col in col_indices if int(col) > 0], reverse=True)
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            return f"Sheet '{sheet_name}' not found in the workbook."
        sheet = workbook[sheet_name]
        for col in col_indices:
            sheet.delete_cols(col)
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width
        for cell in sheet[1]:
            cell.font = cell.font.copy(bold=True)
            cell.alignment = cell.alignment.copy(horizontal='center')
        for row in sheet.iter_rows(min_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            if all(cell.value is None for cell in row):
                sheet.delete_rows(row[0].row)
        workbook.save(file_path)
        workbook.close()
        return f"Columns {', '.join(map(str, col_indices))} have been removed from '{sheet_name}' and formatting applied."
    
    def excel_remove_multiple_rows(self, file_path, sheet_name, row_indices):
        row_indices = sorted([int(row) for row in row_indices if int(row) > 0], reverse=True)
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            return f"Sheet '{sheet_name}' not found in the workbook."
        sheet = workbook[sheet_name]
        for row in row_indices:
            sheet.delete_rows(row)
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width
        for cell in sheet[1]:
            cell.font = cell.font.copy(bold=True)
            cell.alignment = cell.alignment.copy(horizontal='center')
        for row in sheet.iter_rows(min_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            if all(cell.value is None for cell in row):
                sheet.delete_rows(row[0].row)
        workbook.save(file_path)
        workbook.close()
        return f"Rows {', '.join(map(str, row_indices))} have been removed from '{sheet_name}' and formatting applied."
        
    def object_stcode_filter(self, file_path):
        df = pd.read_excel(file_path, dtype=str)
        filtered_df = df[df["Object"].str.contains("S_TCODE", na=False)]
        print(filtered_df)
        filtered_df.to_excel(file_path, index=False, engine='openpyxl')
        print("Filtered rows saved successfully.")

#######################RPA_ROBOT###########################################
    
    def inbounddelivery_json(self, file_path):
        deliveries = defaultdict(list)
        current_delivery = None
        with open(file_path, "r") as file:
            for line in file:
                line = line.strip()
                if line.startswith("InboundDelivery:"):
                    current_delivery = line.split(":")[1]
                elif line.startswith("Item:"):
                    item = {
                        "Item": line.split(":")[1],
                        "Material": next(file).strip().split(":")[1],
                        "DeliveryQuantity": next(file).strip().split(":")[1],
                        "SU": next(file).strip().split(":")[1]
                    }
                    deliveries[current_delivery].append(item)
        json_output = [{"InboundDelivery": key, "TotalItems": value} for key, value in deliveries.items()]
        return json.dumps(json_output, indent=2)
    
    def binallocation_json(self, file_path):
        with open(file_path, "r") as file:
            lines = file.readlines()
        data = []
        current_delivery = None
        for line in lines:
            line = line.strip()
            if line.startswith("InboundDelivery:"):
                if current_delivery:
                    data.append(current_delivery)
                current_delivery = {
                    "InboundDelivery": line.split(":")[1],
                    "VehicleNumber": "",
                    "WarehouseTasks": []
                }
            elif line.startswith("VehicleNumber:"):
                current_delivery["VehicleNumber"] = line.split(":")[1]
            elif line.startswith("WarehousetaskNumber:"):
                task = {
                    "WarehouseTaskNumber": line.split(":")[1],
                    "SourceBin": "",
                    "DestBin": ""
                }
            elif line.startswith("Sourcebin:"):
                task["SourceBin"] = line.split(":")[1]
            elif line.startswith("Destbin:"):
                task["DestBin"] = line.split(":")[1]
                current_delivery["WarehouseTasks"].append(task)
        if current_delivery:
            data.append(current_delivery)
        return json.dumps(data, indent=2)
    

    def generate_list_to_json(self, numbers):
        return json.dumps({f"Delivery_{i}": num for i, num in enumerate(numbers)}, indent=2)
    
    def parse_txt_to_json(self,file_path):
        deliveries = []
        current_delivery = None

        with open(file_path, "r", encoding="utf-8") as file:
            for line in file:
                line = line.strip()
                if not line:
                    continue
                key, value = line.split(":", 1)
                if key == "InboundDelivery":
                    if current_delivery:
                        deliveries.append(current_delivery)
                    current_delivery = {
                        "InboundDelivery": value.strip(),
                        "Material": None,
                        "DeliveryQuantity": None,
                        "Tasks": []
                    }
                elif key == "Material":
                    current_delivery["Material"] = value.strip()
                elif key == "DeliveryQuantity":
                    current_delivery["DeliveryQuantity"] = int(value.strip())
                elif key == "WarehousetaskNumber":
                    task = {"WarehousetaskNumber": int(value.strip())}
                    current_delivery["Tasks"].append(task)
                elif key == "Sourcebin":
                    current_delivery["Tasks"][-1]["Sourcebin"] = value.strip()
                elif key == "Destbin":
                    current_delivery["Tasks"][-1]["Destbin"] = value.strip()
            if current_delivery:
                deliveries.append(current_delivery)
        return json.dumps(deliveries, indent=4)



    def move_columns(self, source_file, source_sheet, source_cols, dest_file, dest_sheet, dest_col):
        src_wb = load_workbook(source_file)
        dest_wb = load_workbook(dest_file)
        if source_sheet not in src_wb.sheetnames:
            return f"Sheet '{source_sheet}' not found in {source_file}"
        if dest_sheet not in dest_wb.sheetnames:
            return f"Sheet '{dest_sheet}' not found in {dest_file}"
        src_ws = src_wb[source_sheet]
        dest_ws = dest_wb[dest_sheet]
        # Ensure columns are integers
        source_cols = sorted([int(col) for col in source_cols])  
        dest_col = int(dest_col)  
        # Find last column in destination sheet
        last_col = dest_ws.max_column
        # Number of columns to be inserted
        num_new_cols = len(source_cols)
        for col in range(last_col, dest_col - 1, -1):
            old_letter = get_column_letter(col)
            new_letter = get_column_letter(col + num_new_cols)
            for row in range(1, dest_ws.max_row + 1):
                dest_ws[f"{new_letter}{row}"].value = dest_ws[f"{old_letter}{row}"].value
        paste_col_index = dest_col
        for src_col in source_cols:
            src_col_letter = get_column_letter(src_col)
            dest_col_letter = get_column_letter(paste_col_index)
            data = [cell.value for cell in src_ws[src_col_letter]]
            for row_idx, value in enumerate(data, start=1):
                dest_ws[f"{dest_col_letter}{row_idx}"].value = value
            src_width = src_ws.column_dimensions[src_col_letter].width
            dest_ws.column_dimensions[dest_col_letter].width = src_width if src_width else 15
            paste_col_index += 1
        src_wb.close()
        dest_wb.save(dest_file)
        dest_wb.close()
        return f"Columns {', '.join([get_column_letter(col) for col in source_cols])} moved to column {get_column_letter(dest_col)} in '{dest_sheet}', shifting existing data to the right."
    
    def get_column_excel_to_txt_create(self, excel_path, txt_path, column_name):
        df = pd.read_excel(excel_path, dtype=str)
        values = df[column_name].dropna().astype(float).astype(int).tolist()
        unique_values = sorted(set(values))
        with open(txt_path, "w") as file:
            file.writelines(f"{item}\n" for item in unique_values)
        print("Saved successfully!")
    
    def compare_vendors_data(self, defaultfile, comparefile, outputfile):
        wb1 = openpyxl.load_workbook(defaultfile)
        wb2 = openpyxl.load_workbook(comparefile)
        sheet1 = wb1.active
        sheet2 = wb2.active
        col_b_sheet1 = 2
        sheet1.insert_cols(col_b_sheet1 + 1)
        sheet1.insert_cols(col_b_sheet1 + 2)  
        bold_font = Font(bold=True)
        sheet1.cell(row=1, column=col_b_sheet1 + 1, value="Supplier Name").font = bold_font
        sheet1.cell(row=1, column=col_b_sheet1 + 2, value="Account Group").font = bold_font
        vendor_map = {}
        for row in sheet2.iter_rows(min_row=2, values_only=True):
            vendor_map[row[0]] = (row[1], row[2])
        for row in range(2, sheet1.max_row + 1):
            vendor = sheet1.cell(row=row, column=col_b_sheet1).value
            if vendor in vendor_map:
                supplier_name, account_group = vendor_map[vendor]
                sheet1.cell(row=row, column=col_b_sheet1 + 1, value=supplier_name)
                sheet1.cell(row=row, column=col_b_sheet1 + 2, value=account_group)
        output_path = outputfile
        wb1.save(output_path)
        print("Excel file saved")


    def excel_remove_except_column(self, file_path, sheet_name, keep_column_names):
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            return f"Sheet '{sheet_name}' not found in the workbook."
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        keep_indices = [headers.index(name) + 1 for name in keep_column_names if name in headers]
        if not keep_indices:
            return "None of the specified columns were found in the sheet."
        col_indices = sorted([col for col in range(1, sheet.max_column + 1) if col not in keep_indices], reverse=True)
        for col in col_indices:
            sheet.delete_cols(col)
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width
        for cell in sheet[1]:
            cell.font = cell.font.copy(bold=True)
            cell.alignment = cell.alignment.copy(horizontal='center')
        for row in reversed(list(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column))):
            if all(cell.value is None for cell in row):
                sheet.delete_rows(row[0].row)
        workbook.save(file_path)
        workbook.close()
        return f"Only columns {', '.join(keep_column_names)} have been kept in '{sheet_name}', and formatting applied."
    
    def swap_and_shift_columns(self, file_path, sheet_name, column_name, current_col_index):

        df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")

        current_col_index = int(current_col_index) - 1

        second_col_name = df.columns[1]
        second_col_data = df.pop(second_col_name)
        column_data = df.pop(column_name)

        df.insert(1, column_name, column_data)
        df.insert(2, second_col_name, second_col_data)

        df.to_excel(file_path, sheet_name=sheet_name, index=False, engine="openpyxl")

    def compare_deptor_data(self, defaultfile ,comparefile ,outputfile):
        wb1 = openpyxl.load_workbook(defaultfile)
        wb2 = openpyxl.load_workbook(comparefile)

        sheet1 = wb1.active
        sheet2 = wb2.active

        col_b_sheet1 = 2
        sheet1.insert_cols(col_b_sheet1 + 1)
        sheet1.insert_cols(col_b_sheet1 + 2)

        bold_font = Font(bold=True)
        sheet1.cell(row=1, column=col_b_sheet1 + 1, value="Name 1").font = bold_font
        sheet1.cell(row=1, column=col_b_sheet1 + 2, value="Account group").font = bold_font

        vendor_map = {}
        for row in sheet2.iter_rows(min_row=2, values_only=True):
            if row[0]:  
                vendor_id = str(row[0]).strip()
                vendor_map[vendor_id] = (row[1], row[2])
        print("🔍 Vendor Map Loaded:", vendor_map)

        for row in range(2, sheet1.max_row + 1):
            vendor = sheet1.cell(row=row, column=col_b_sheet1).value
            vendor = str(vendor).strip() if vendor else None 
            if vendor in vendor_map:
                name_1, account_group = vendor_map[vendor]
                sheet1.cell(row=row, column=col_b_sheet1 + 1, value=name_1)
                sheet1.cell(row=row, column=col_b_sheet1 + 2, value=account_group)
            else:
                print(f"Vendor ID {vendor} not found in sheet2")       
        wb1.save(outputfile)



    def get_last_day_of_month(self):
        year = datetime.now().year
        month = datetime.now().month
        last_day = calendar.monthrange(year, month)[1]
        return f"{last_day:02d}.{month:02d}.{year}"

    def get_start_day_of_month(self):
        year = datetime.now().year
        month = datetime.now().month
        return f"01.{month:02d}.{year}"
    
    def split_numeric_alphabetic(self,text):
        match = re.match(r'(\d+)\s*(\D+)', text)
        if match:
            return match.group(1), match.group(2).strip()
        return None, None
    def rpa_report(self, file_path, output_file):
        def parse_datetime(date, time):
            return datetime.strptime(f"{date} {time}", "%d.%m.%Y %H:%M:%S")
        with open(file_path, "r", encoding="utf-8") as file:
            content = file.read()
        
        data = {
            "vehicle_entry": re.search(r"Vehicle Entry :Date:(\d{2}\.\d{2}\.\d{4}) Time:(\d{2}:\d{2}:\d{2})", content),
            "goods_receipt": re.search(r"Goods Receipt Created :Date:(\d{2}\.\d{2}\.\d{4}) Time:(\d{2}:\d{2}:\d{2})", content),
            "storage_location": re.search(r"Material placed in Storage Location Confirmed :Date:(\d{2}\.\d{2}.\d{4}) Time:(\d{2}:\d{2}:\d{2})", content),
            "vendor_number": re.search(r"Vendor Number :(\d+)", content),
            "vendor_name": re.search(r"Vendor Name :([\w\s]+)", content),
            "inbound_delivery": re.search(r"Inbound Delivery Number :(\d+)", content),
            "vehicle_number": re.search(r"Vehicle Number :([\w\s]+)", content),
            "delivery_date": re.search(r"Delivery Date :(\d{2}\.\d{2}\.\d{4})", content),
            "purchase_order": re.search(r"Purchase Order :(\d+)", content),
            "material_number": re.findall(r"Material Number :(\d+)", content),
            "quantity": re.findall(r"Quanitity :(\d+)", content),
            "unit": re.findall(r"Unit :(\w+)", content),
            "material_desc": re.findall(r"Material Description :(.*)", content)
        }
        
        data = {key: match.groups() if isinstance(match, re.Match) else match for key, match in data.items()}
        timestamps = {}
        if data["vehicle_entry"]:
            timestamps["vehicle_entry"] = parse_datetime(*data["vehicle_entry"])
        if data["goods_receipt"]:
            timestamps["goods_receipt"] = parse_datetime(*data["goods_receipt"])
        if data["storage_location"]:
            timestamps["storage_location"] = parse_datetime(*data["storage_location"])

        # Compute time differences
        time_differences = {}
        if "vehicle_entry" in timestamps and "goods_receipt" in timestamps:
            time_differences["entry_to_receipt"] = timestamps["goods_receipt"] - timestamps["vehicle_entry"]
        if "goods_receipt" in timestamps and "storage_location" in timestamps:
            time_differences["receipt_to_storage"] = timestamps["storage_location"] - timestamps["goods_receipt"]
        if "vehicle_entry" in timestamps and "storage_location" in timestamps:
            time_differences["total"] = timestamps["storage_location"] - timestamps["vehicle_entry"]
        data["time_differences"] = time_differences

        line_items = "".join(
            f"<tr><td>{i+1}</td><td>{mat[0]}</td><td>{mat[1]}</td><td>{mat[2]}</td><td>{mat[3]}</td><td>Completed</td></tr>"
            for i, mat in enumerate(zip(data['material_number'], data['quantity'], data['unit'], data['material_desc']))
        )

        timeline_items = [
            f"<li class='timeline-item completed'>Vehicle Entry - <b>Date:</b> {data['vehicle_entry'][0]} <b>Time:</b> {data['vehicle_entry'][1]}</li>"
            if data.get('vehicle_entry') else "<li class='timeline-item pending'>Vehicle Entry - <b>Date:</b> - <b>Time:</b> 00:00:00</li>",

            f"<li class='timeline-item completed'>Goods Receipt Created - <b>Date:</b> {data['goods_receipt'][0]} <b>Time:</b> {data['goods_receipt'][1]}</li>"
            if data.get('goods_receipt') else "<li class='timeline-item pending'>Goods Receipt Created - <b>Date:</b> - <b>Time:</b> 00:00:00</li>",

            f"<li class='timeline-item completed'>Material placed in Storage Location - <b>Date:</b> {data['storage_location'][0]} <b>Time:</b> {data['storage_location'][1]}</li>"
            if data.get('storage_location') else "<li class='timeline-item pending'>Material placed in Storage Location - <b>Date:</b> - <b>Time:</b> 00:00:00</li>"
        ]

        total_time = time_differences.get("total", "N/A")
        timeline_items.append(
            f"<li class='timeline-item summary' style='font-weight: bold; color: green;'>Total Time Taken: <b>{total_time}</b> [Vehicle Entry → Material Placed in Storage Location]</li>"
        )

        html_content = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Timeline Tracker</title>
            <style>
                body {{ font-family: Arial, sans-serif; max-width: 900px; margin: auto; padding: 20px; background-color: #f6f6f6; }}
                .header {{ display: flex; align-items: center; gap: 100px; background-color: #0073e6; padding: 15px; border-radius: 8px; color: white; font-size: 20px; font-weight: bold; }}
                .header img {{ height: 60px; }}
                .table-container {{ margin-top: 30px; border: 2px solid black; border-radius: 5px; background-color: white; padding: 10px; }}
                table {{ width: 100%; border-collapse: collapse; text-align: center; }}
                th, td {{ padding: 10px; border: 1px solid black; }}
                th {{ background-color: yellow; }}
                .timeline-container {{ margin-top: 20px; background-color: white; padding: 20px; border-radius: 5px; border: 2px solid black; }}
                .timeline {{ padding-left: 30px; list-style: none; }}
                .timeline-item {{ padding: 15px; background: #f6f6f6; margin-bottom: 10px; box-shadow: 2px 2px 10px rgba(0, 0, 0, 0.1); }}
                .completed {{ color: green; font-weight: bold; }}
                .pending {{ color: gray; font-weight: bold; }}
                .timeline::before {{
                        content: "";
                        position: absolute;
                        left: 12px;
                        top: 0;
                        height: 100%;
                        width: 4px;
                    }}
                    .timeline-item {{
                        position: relative;
                        padding: 15px 0;
                        padding-left: 40px;
                    }}
                    .timeline-item::before {{
                        content: "✔️";
                        position: absolute;
                        left: 0;
                        top: 50%;
                        transform: translateY(-50%);
                        font-size: 20px;
                        color: green;
                    }}

                    .pending::before {{
                        content: "🔴";
                        color: red;
                    }}
            </style>
        </head>
        <body>
            <div class="header">
                <img src="https://symphony4cloud.com/static/media/logo-1.a84c61ed3670ed0777c8.png" alt="Logo">
                <div class="title">Symphony AI Powered Automation Report</div>
            </div>
            <div class="table-container">
                <table>
                    <tr><th colspan="6">Header</th></tr>
                    <tr><th>Vendor Number</th><th>Vendor Name</th><th>Purchase Order</th><th>Inbound Delivery</th><th>Vehicle Number</th><th>Delivery Date</th></tr>
                    <tr>
                        <td>{data['vendor_number'][0] if data['vendor_number'] else ''}</td>
                        <td>{data['vendor_name'][0] if data['vendor_name'] else ''}</td>
                        <td>{data['purchase_order'][0] if data['purchase_order'] else ''}</td>
                        <td>{data['inbound_delivery'][0] if data['inbound_delivery'] else ''}</td>
                        <td>{data['vehicle_number'][0] if data['vehicle_number'] else ''}</td>
                        <td>{data['delivery_date'][0] if data['delivery_date'] else ''}</td>
                    </tr>
                    <tr><th>S.No</th><th>Material Number</th><th>Quantity</th><th>Unit</th><th>Material Description</th><th>Delivery Status</th></tr>
                        {line_items}
                </table>
            </div>
            <div class="timeline-container">
                <ul class="timeline">
                    {"".join(timeline_items)}
                </ul>
            </div>
        </body>
        </html>
        """
        with open(output_file, "w", encoding="utf-8") as file:
            file.write(html_content)

    def search_transport_id(self, json_file, transport_id):
        try:
            with open(json_file, 'r', encoding='utf-8') as file:
                data = json.load(file)
        #try founded vehicle
            if isinstance(data, list):
                for item in data:
                    if item.get("Means of Trans. ID") == transport_id:
                        return item.get("Delivery"), {
                            "Message": "Match found",
                            "Delivery": item.get("Delivery"),
                            "Deliv. date(From/to)": item.get("Deliv. date(From/to)")
                        }
        #catch not found vehioc;le
            elif isinstance(data, dict):
                if data.get("Means of Trans. ID") == transport_id:
                    return item.get("Delivery"), {
                        "Message": "Match found",
                        "Delivery": data.get("Delivery"),
                        "Deliv. date(From/to)": data.get("Deliv. date(From/to)")
                    }
            return f"No match found for this vehicle {transport_id} in current month"
        except Exception as e:
            return f"Error: {e}"
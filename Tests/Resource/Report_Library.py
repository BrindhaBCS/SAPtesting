import pythoncom
import win32com.client
import time
from pythoncom import com_error
import robot.libraries.Screenshot as screenshot
import os
from robot.api import logger
import sys
import ast
import re
import glob
import  json
# from docx2pdf import convert
from PIL import Image
# import docx
# from docx import Document
# from docx.enum.table import WD_TABLE_ALIGNMENT
# from docx.shared import Cm, Pt, Mm, Inches
# from docx.enum.section import WD_ORIENT
from openpyxl import load_workbook
import pandas as pd
import  openpyxl
import json
from fpdf import FPDF
from PIL import Image
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC 
from openpyxl import load_workbook 
import shutil
import pandas as pd
import datetime
import os
import csv
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException


class Report_Library:    
    def __init__(self):
        self.name = "DefaultReport"
    def copy_images(self, source_dir, target_dir):
        # Ensure the target directory exists
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)

        # Supported image formats by PIL (Pillow)
        image_formats = ('.jpeg', '.jpg', '.png', '.gif', '.bmp', '.tiff', '.webp')

        # Iterate over files in the source directory
        for file_name in os.listdir(source_dir):
            file_path = os.path.join(source_dir, file_name)

            # Check if it's a file and if it has a valid image format
            if os.path.isfile(file_path):
                try:
                    with Image.open(file_path) as img:  # This will fail if the file is not a valid image
                        if file_name.lower().endswith(image_formats):
                            target_path = os.path.join(target_dir, file_name)
                            shutil.copy(file_path, target_path)
                            print(f"Copied: {file_name}")
                except Exception as e:
                    print(f"Skipped: {file_name} - Not a valid image file or format not supported. Error: {e}")
    def remove_rows_before_start_row(self, file_path, sheet_name, start_row):
        start_row = int(start_row)

        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        for _ in range(start_row - 1):  # Start row number is 1-based
            sheet.delete_rows(1)  # Always delete the first row

        workbook.save(file_path)
        workbook.close()
        return f"All rows before row {start_row} have been removed."

    def clean_excel(self, file_path, sheet_name):
        try:
            workbook = load_workbook(file_path)
            if sheet_name not in workbook.sheetnames:
                raise Exception(f"Sheet '{sheet_name}' does not exist in the file.")

            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = cell.value.strip()  # Remove leading and trailing whitespace

            # Remove completely empty rows (backward iteration to avoid index shift)
            for row_idx in range(sheet.max_row, 0, -1):
                if all(sheet.cell(row=row_idx, column=col_idx).value in [None, ""] for col_idx in range(1, sheet.max_column + 1)):
                    sheet.delete_rows(row_idx)

            # Remove completely empty columns (backward iteration to avoid index shift)
            for col_idx in range(sheet.max_column, 0, -1):
                if all(sheet.cell(row=row_idx, column=col_idx).value in [None, ""] for row_idx in range(1, sheet.max_row + 1)):
                    sheet.delete_cols(col_idx)

            # Save changes back to the file
            workbook.save(file_path)
            print("\033[92m❗ Excel sheet cleaned successfully and saved at:", file_path)  # Green exclamation mark

        except Exception as e:
            print("\033[92m❗ Failed to clean Excel sheet:", str(e))  # Green exclamation mark

    def fiori_html_report(self, excel_file, html_file, highlight_text):
        df = pd.read_excel(excel_file)

        # Initialize counters
        pass_count = 0
        warn_count = 0
        for index, row in df.iterrows():
            row_text = " ".join(row.astype(str).tolist())
            if 'Authorization check successful' in row_text:
                pass_count += 1
            if 'No authorization in user master record' in row_text:
                warn_count += 1

        current_time = datetime.datetime.utcnow().strftime('%d-%m-%Y Time: %H:%M:%S UTC')

        #HTML
        html_content = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Fiori LaunchPad Report</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 0;
                    background-color: #f4f4f4;
                }}
                .container {{
                    width: 90%;
                    margin: 30px auto;
                    padding: 20px;
                    background-color: #ffffff;
                    border-radius: 8px;
                    box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
                }}
                h1 {{
                    color: #1f3a64;
                    text-align: center;
                }}
                .count-box {{
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    gap: 20px;
                    margin: 20px 0;
                }}
                .count-item {{
                    display: flex;
                    align-items: center;
                    gap: 10px;
                    font-size: 18px;
                }}
                .pass-circle {{
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    background-color: green;
                    border-radius: 50%;
                    margin-right: 8px; /* Adds spacing between the circle and text */
                }}
                .warn-circle {{
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    background-color: red;
                    border-radius: 50%;
                    margin-right: 8px; /* Adds spacing between the circle and text */
                }}

                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th, td {{
                    padding: 12px;
                    text-align: left;
                    border: 1px solid #ddd;
                }}
                th {{
                    background-color: #1f3a64;
                    color: white;
                }}
                td {{
                    background-color: #fafafa;
                }}
                table tr:nth-child(even) td {{
                    background-color: #f2f2f2;
                }}
                table tr:hover td {{
                    background-color: #e9e9e9;
                }}
                .row-green {{
                    background-color: #d4edda;
                    color: green;
                }}
                .row-red {{
                    background-color: #ff0000 ;
                    color:red ;
                }}
            </style>
        </head>
        <body>

        <div class="container">
            <h1>Fiori LaunchPad Report</h1>
                <div class="count-box">
                    <div class="count-item">
                        <span class="pass-circle"></span> Authorization Count: <span>{pass_count}</span>
                    </div>
                    <div class="count-item">
                        <span class="warn-circle"></span> No Authorization Count: <span>{warn_count}</span>
                    </div>
                    <div class="count-item">
                        Generated At: <span>{current_time}</span>
                    </div>
                </div>
            <table>
                <thead>
                    <tr>
        """

        # Add table headers (columns from DataFrame)
        headers = df.columns.tolist()
        for header in headers:
            html_content += f"<th>{header}</th>"

        html_content += """
                    </tr>
                </thead>
                <tbody>
        """

        # Add table rows (data from DataFrame)
        for index, row in df.iterrows():
            # Check if the target text is in any cell of the row
            if row.astype(str).str.contains(highlight_text, case=False).any():
                html_content += '<tr class="row-red">'  
            else:
                html_content += '<tr class="row-green">'  

            for value in row:
                html_content += f"<td>{value}</td>"

            html_content += "</tr>"

        html_content += """
                </tbody>
            </table>
        </div>

        </body>
        </html>
        """

        with open(html_file, "w") as file:
            file.write(html_content)

        print(f"HTML report created successfully: {html_file}")

    def sarole_html_report(self, excel_file, html_file, highlight_text):
        df = pd.read_excel(excel_file)

        # Initialize counters
        pass_count = 0
        warn_count = 0
        for index, row in df.iterrows():
            row_text = " ".join(row.astype(str).tolist())
            if 'Authorization check successful' in row_text:
                pass_count += 1
            if 'No authorization in user master record' in row_text:
                warn_count += 1

        current_time = datetime.datetime.utcnow().strftime('%d-%m-%Y Time: %H:%M:%S UTC')

        #HTML
        html_content = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Security Authorization Report</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 0;
                    background-color: #f4f4f4;
                }}
                .container {{
                    width: 90%;
                    margin: 30px auto;
                    padding: 20px;
                    background-color: #ffffff;
                    border-radius: 8px;
                    box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
                }}
                h1 {{
                    color: #1f3a64;
                    text-align: center;
                }}
                .count-box {{
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    gap: 20px;
                    margin: 20px 0;
                }}
                .count-item {{
                    display: flex;
                    align-items: center;
                    gap: 10px;
                    font-size: 18px;
                }}
                .pass-circle {{
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    background-color: green;
                    border-radius: 50%;
                    margin-right: 8px; /* Adds spacing between the circle and text */
                }}
                .warn-circle {{
                    display: inline-block;
                    width: 15px;
                    height: 15px;
                    background-color: red;
                    border-radius: 50%;
                    margin-right: 8px; /* Adds spacing between the circle and text */
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th, td {{
                    padding: 12px;
                    text-align: left;
                    border: 1px solid #ddd;
                }}
                th {{
                    background-color: #1f3a64;
                    color: white;
                }}
                td {{
                    background-color: #fafafa;
                }}
                table tr:nth-child(even) td {{
                    background-color: #f2f2f2;
                }}
                table tr:hover td {{
                    background-color: #e9e9e9;
                }}
                .row-green {{
                    background-color: #d4edda;
                    color: green;
                }}
                .row-red {{
                    background-color: #ff0000 ;
                    color:red ;
                }}
            </style>
        </head>
        <body>
        <div class="container">
            <h1>Security Authorization Report</h1>
                <div class="count-box">
                    <div class="count-item">
                        <span class="pass-circle"></span> Authorization Count: <span>{pass_count}</span>
                    </div>
                    <div class="count-item">
                        <span class="warn-circle"></span> No Authorization Count: <span>{warn_count}</span>
                    </div>
                    <div class="count-item">
                        Generated At: <span>{current_time}</span>
                    </div>
                </div>
            <table>
                <thead>
                    <tr>
        """
        # Add table headers (columns from DataFrame)
        headers = df.columns.tolist()
        for header in headers:
            html_content += f"<th>{header}</th>"

        html_content += """
                    </tr>
                </thead>
                <tbody>
        """
        # Add table rows (data from DataFrame)
        for index, row in df.iterrows():
            # Check if the target text is in any cell of the row
            if row.astype(str).str.contains(highlight_text, case=False).any():
                html_content += '<tr class="row-red">'  
            else:
                html_content += '<tr class="row-green">'  
            for value in row:
                html_content += f"<td>{value}</td>"
            html_content += "</tr>"
        html_content += """
                </tbody>
            </table>
        </div>
        </body>
        </html>
        """
        with open(html_file, "w") as file:
            file.write(html_content)
        print(f"HTML report created successfully: {html_file}")


    def convert_csv_to_excel(self, csv_file_path, excel_file_path):
        try:
            # Read the CSV file with explicit delimiter as semicolon and handle the BOM
            df = pd.read_csv(csv_file_path, delimiter=';', decimal=',', encoding='ISO-8859-1', engine='python', on_bad_lines='skip', skiprows=1)
            # Check if data is loaded correctly
            print(df.head())  # Preview data before saving
            # Save as Excel file
            df.to_excel(excel_file_path, index=False)
            print(f"CSV file successfully converted to Excel: {excel_file_path}")
        except Exception as e:
            print(f"Error during conversion: {e}")


    def cop_webexcel_scrap(self, file_path, output_file):
        # Create a new Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        with open(file_path, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file, delimiter=';')
            row_number = 1
            for row in reader:
                if len(row) < 10:
                    continue
                snote_value = row[1]
                title = row[2]
                cvss_score = row[3]
                cvss_vector = row[4]
                cvss_components = cvss_vector.split('/')
                if len(cvss_components) >= 6:
                    extracted_cvss = '/'.join(cvss_components[-3:])
                else:
                    extracted_cvss = ""
                category = row[5]
                priority = row[6]
                released = row[7]
                first_released_on = row[8]
                link = row[9]
                # Write data to the Excel sheet
                ws[f'A{row_number}'] = snote_value
                ws[f'B{row_number}'] = title
                ws[f'C{row_number}'] = cvss_score
                ws[f'D{row_number}'] = extracted_cvss
                ws[f'E{row_number}'] = category
                ws[f'F{row_number}'] = priority
                ws[f'G{row_number}'] = released
                ws[f'H{row_number}'] = first_released_on
                ws[f'I{row_number}'] = link
                row_number += 1
                #heder
                ws['A1'] = "Web Data"
                ws['B1'] = "Title"
                ws['C1'] = "CVSS Score"
                ws['D1'] = "CVSS Vector"
                ws['E1'] = "Category"
                ws['F1'] = "Priority"
                ws['G1'] = "Released On"
                ws['H1'] = "First Released On"
                ws['I1'] = "Link"

        # Save the workbook
        wb.save(output_file)
        print(f"Extracted data has been saved to {output_file}")


    def cop_sapexcel_scrap(self, input_file, output_file):
        df = pd.read_excel(input_file)
        wb = load_workbook(output_file)
        ws = wb.active 
        ws['J1'] = "NUMM"
        ws['K1'] = "PRSTATUS"
        row_number = 2
        for index, row in df.iterrows():
            col2_data = row[0]  # Column 2 (index 1 )
            col9_data = row[2]  # Column 9 (index 8)
            if pd.isna(col2_data) or pd.isna(col9_data) or "NUMM" in str(col2_data) or "PRSTATUS" in str(col2_data) or str(col2_data).strip() == '' or str(col9_data).strip() == '':
                continue
            col2_data = str(col2_data).lstrip('0')
            if col2_data.startswith('000'):
                col2_data = col2_data[3:]
            col9_data = str(col9_data).lstrip('0')
            if col9_data.startswith('000'):
                col9_data = col9_data[3:]
            ws[f'J{row_number}'] = col2_data
            ws[f'K{row_number}'] = col9_data
            row_number += 1
        wb.save(output_file)
        print(f"Data from columns 2 and 9 of {input_file} (with first three zeros removed) has been saved to columns 2 and 3 of {output_file}")


    def cop_excel_compare(self, input_file, input_sheet_name, output_file):
        # Read the Excel file
        df = pd.read_excel(input_file, sheet_name=input_sheet_name, dtype=str)   
        col_a = df.iloc[:, 0]  # Column A
        col_b = df.iloc[:, 1]  # Column B
        col_c = df.iloc[:, 2]  # Column C
        col_d = df.iloc[:, 3]  # Column D
        col_e = df.iloc[:, 4]  # Column E
        col_f = df.iloc[:, 5]  # Column F
        col_g = pd.to_datetime(df.iloc[:, 6], errors='coerce').dt.strftime('%d-%m-%Y')
        col_h = df.iloc[:, 7]  # Column H
        col_i = df.iloc[:, 8]  # Column I
        col_j = df.iloc[:, 9]  # Column J
        col_k = df.iloc[:, 10] # Column K

        def map_status(prstatus):
            if prstatus == '-':
                return "Cannot be implemented"
            elif prstatus in ['U', 'u']:
                return "Incompletely implemented"
            elif prstatus in ['E', 'e']:
                return "Completely implemented"
            elif prstatus in ['N', 'n', None, '']:
                return "Not yet implemented"
            elif prstatus == "Unknown Status": 
                return "Not yet implemented"
            else:
                return "Invalid Status"

        comparison_results = []
        unmatched_results = []

        for i in range(len(col_a)):
            matched = False
            for j in range(len(col_j)):
                if col_a[i] == col_j[j]:  # Match found
                    comparison_results.append({
                        "Web Data": col_a[i],    
                        "Title": col_b[i],    
                        "CVSS Score": col_c[i],    
                        "CVSS Vector": col_d[i],    
                        "Category": col_e[i],
                        "Priority": col_f[i],
                        "Released On": col_g[i], 
                        "First Released On": col_h[i],
                        "Link": col_i[i],
                        "SAP NUMM": col_j[j],  
                        "SAP PRSTATUS": col_k[j],  
                        "Status": map_status(col_k[j])
                    })
                    matched = True
            if not matched:
                unmatched_results.append({
                    "Web Data": col_a[i],    
                    "Title": col_b[i],    
                    "CVSS Score": col_c[i],    
                    "CVSS Vector": col_d[i],    
                    "Category": col_e[i],
                    "Priority": col_f[i],
                    "Released On": col_g[i],  
                    "First Released On": col_h[i],
                    "Link": col_i[i],
                    "SAP NUMM": "Unmatched",  
                    "SAP PRSTATUS": "Unmatched",  
                    "Status": "Not yet implemented"
                })

        # Sort
        comparison_results = sorted(comparison_results, key=lambda x: pd.to_datetime(x["Released On"], format='%d-%m-%Y', errors='coerce') if pd.notna(x["Released On"]) else pd.Timestamp.min)
        unmatched_results = sorted(unmatched_results, key=lambda x: pd.to_datetime(x["Released On"], format='%d-%m-%Y', errors='coerce') if pd.notna(x["Released On"]) else pd.Timestamp.min)
        comparison_df = pd.DataFrame(comparison_results)
        unmatched_df = pd.DataFrame(unmatched_results)
        final_df = pd.concat([comparison_df, unmatched_df], ignore_index=True)
        final_df["Released On"] = final_df["Released On"].astype(str).replace("NaT", "")
        try:
            wb = load_workbook(output_file)
        except (FileNotFoundError, InvalidFileException):
            wb = load_workbook(input_file)

        if "Master Data" in wb.sheetnames:
            del wb["Master Data"]
        ws = wb.create_sheet(title="Master Data", index=3)
        ws.append(list(final_df.columns))
        for row in final_df.itertuples(index=False, name=None):
            ws.append(row)
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Blue
        red_fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")   # Red
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") # Green

        for row in range(2, len(final_df) + 2):
            prstatus_value = ws[f'K{row}'].value
            status_value = ws[f'L{row}'].value

            if prstatus_value == '-':
                ws[f'K{row}'].fill = blue_fill  
            elif prstatus_value in ['U', 'u']:
                ws[f'K{row}'].fill = red_fill  
            elif prstatus_value in ['E', 'e']:
                ws[f'K{row}'].fill = green_fill  
            elif prstatus_value in ['N', 'n', None, '']:
                ws[f'K{row}'].fill = red_fill 
            elif prstatus_value == 'Unmatched':
                ws[f'K{row}'].fill = red_fill 

            if status_value == "Not yet implemented":
                ws[f'L{row}'].fill = red_fill  
            elif status_value == "Cannot be implemented":
                ws[f'L{row}'].fill = blue_fill
            elif status_value == "Completely implemented":
                ws[f'L{row}'].fill = green_fill
            elif status_value == "Incompletely implemented":
                ws[f'L{row}'].fill = red_fill
            elif status_value is None:
                ws[f'L{row}'].fill = red_fill

        # Save Excel file
        wb.save(output_file)
        print(f"Updated file with sorted matched {output_file}")
    def basis_formate_json_data(self, input_string, removed_lines):
        def remove_first_n_lines(input_string, removed_lines):
            lines = input_string.split('\n')
            print('\n'.join(lines[removed_lines:]).strip())

        def extract_json_data(data):
            start_index = data.find("[")
            if start_index == -1:
                print({"status": "error", "message": "No JSON data found"})
            json_data = data[start_index:]

            try:
                parsed_json = json.loads(json_data)  # Convert string to Python JSON object
                print(parsed_json)
            except json.JSONDecodeError as e:
                print({"status": "error", "message": f"JSON Decode Error: {str(e)}"})
        cleaned_string = remove_first_n_lines(input_string, removed_lines)
        json_result = extract_json_data(cleaned_string)
        print (json_result)
        return  json_result